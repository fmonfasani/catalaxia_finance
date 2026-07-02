"""
GRIM (Grimoldi) - COMPLETE PIPELINE from CNV AIF2 discovery to dashboard
"""
import csv, re, sqlite3, json, math
from pathlib import Path
from datetime import datetime
import requests, html as ihtml
import yfinance as yf

BASE = Path(r'D:\Software Development\Porfolio\catalaxia-cedears-prod')
H = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120', 'Accept-Language': 'es-AR'}

CODIGOS = {
    "1122500": "Cash", "1121999": "Receivables", "1120100": "Inventory",
    "1139999": "AssetsCurrent", "1110100": "PPE", "1110200": "Intangibles",
    "1119999": "AssetsNonCurrent", "1999999": "Assets",
    "2210999": "Capital", "2211999": "Reservas", "2212999": "ResultadosNoAsignados",
    "2299999": "Equity",
    "2322200": "DebtCurrent", "2321999": "Payables", "2339999": "LiabilitiesCurrent",
    "2312300": "DebtNonCurrent", "2319999": "LiabilitiesNonCurrent", "2399999": "Liabilities",
    "3000100": "Revenue", "3000200": "COGS", "3009999": "GrossProfit",
    "3011600": "DA", "3019999": "OperatingIncome", "3021400": "IngresosFinancieros",
    "3021500": "InterestExpense", "3021800": "RECPAM", "3029999": "PretaxIncome",
    "3031100": "IncomeTax", "3049999": "NetIncome", "3099999": "ResultadoIntegral",
    "3240000": "CashFlowNeto",
    "8000000": "EPS_basico", "8000001": "EPS_diluido", "8000003": "EBIT",
    "8000004": "EBITDA", "8000005": "WorkingCapital",
}

def parse(html_text):
    txt = ihtml.unescape(re.sub("<[^>]+>", " ", html_text))
    txt = re.sub(r"\s+", " ", txt)
    pares = {}
    for m in re.finditer(r"([1-8]\d{6})\s+[A-ZÁÉÍÓÚÑ()/.,\s]+?\s+(-?\d+\.\d{2})", txt):
        pares.setdefault(m.group(1), float(m.group(2)))
    return {CODIGOS[c]: v for c, v in pares.items() if c in CODIGOS}

def detect_year(html_text):
    """Detect fiscal year from HTML title or content."""
    m = re.search(r'(?:al|del|PER[IÍ]ODO)\s+(\d{2}/\d{2}/\d{4})', html_text[:5000], re.I)
    if m:
        return m.group(1)
    m = re.search(r'(\d{4})', html_text[:2000])
    return m.group(1) if m else "unknown"

# ── FASE 1: Discover all NIIF GUIDs for GRIM ──
print("=" * 70)
print("GRIM (Grimoldi SA) - PIPELINE COMPLETO")
print("=" * 70)

refined = BASE / 'scripts' / 'tickets' / 'cnv' / 'datos' / 'links_eeff_refined.csv'
grim_niif = []
with open(refined, encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        if row['ticker'].upper() == 'GRIM' and row['formTypeId'] == '147':
            grim_niif.append(row)

# Sort by presentationId descending
grim_niif.sort(key=lambda x: int(float(x['presentationId'])), reverse=True)
print(f"\nFASE 1: {len(grim_niif)} presentaciones NIIF encontradas")

# ── FASE 2: Extract financial data for each year ──
print(f"\nFASE 2: Extrayendo datos financieros...")
all_years = {}
for row in grim_niif:
    pid = int(float(row['presentationId']))
    guid = row['guid']
    try:
        r = requests.get(row['url'], headers=H, timeout=25, verify=False)
        datos = parse(r.text)
        if 'Assets' in datos and 'Equity' in datos:
            year = detect_year(r.text)
            all_years[pid] = {'year': year, 'datos': datos, 'guid': guid}
            print(f"  PID={pid} Year={year} OK: Assets={datos['Assets']:,.0f} Equity={datos.get('Equity',0):,.0f} NI={datos.get('NetIncome',0):,.0f}")
        if len(all_years) >= 5:  # Get up to 5 years
            break
    except Exception as e:
        print(f"  PID={pid} ERROR: {str(e)[:40]}")

# Sort years by PID (ascending)
sorted_pids = sorted(all_years.keys())
data_by_year = {all_years[p]['year']: all_years[p]['datos'] for p in sorted_pids}

print(f"\n  Total años con datos: {len(data_by_year)}")
for yr, d in data_by_year.items():
    ident = abs(d.get('Liabilities',0) + d.get('Equity',0) - d.get('Assets',1)) / abs(d.get('Assets',1)) * 100
    print(f"    {yr}: Rev={d.get('Revenue',0):,.0f} NI={d.get('NetIncome',0):,.0f} Eq={d.get('Equity',0):,.0f} DebtC={d.get('DebtCurrent',0):,.0f} EBITDA={d.get('EBITDA',0):,.0f} Ident={ident:.2f}%")

# ── FASE 3: Get price data ──
print(f"\nFASE 3: Precio BYMA")
t = yf.Ticker("GRIM.BA")
info = t.info
price_ars = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
year_high = info.get('fiftyTwoWeekHigh')
year_low = info.get('fiftyTwoWeekLow')
shares = info.get('sharesOutstanding')
market_cap_ars = info.get('marketCap')
if not year_high or not year_low:
    hist = t.history(period='1y')
    if len(hist) > 0:
        year_high = year_high or float(hist['High'].max())
        year_low = year_low or float(hist['Low'].min())

# ARS/USD rate
try:
    fx_r = requests.get('https://api.bluelytics.com.ar/v2/latest', timeout=10)
    fx = float(fx_r.json().get('blue', {}).get('value_sell', 1200))
except:
    fx = 1200

price_usd = price_ars / fx if price_ars else None
print(f"  Precio: ARS${price_ars:,.2f} = USD${price_usd:.2f} (ARS/USD={fx})")
print(f"  52w High: {year_high}  Low: {year_low}")
print(f"  Shares: {shares:,}  MarketCap: {market_cap_ars:,.0f}")

# ── FASE 4: Calculate all 13 ratios ──
print(f"\nFASE 4: Calculando 13 ratios...")

def safe_div(a, b):
    return a / b if (a is not None and b not in (None, 0)) else None

latest = list(data_by_year.values())[-1]  # Last sorted = most recent
latest_yr = list(data_by_year.keys())[-1]

R = {}  # ratios dict

# 1. Precio en u$s
R['Precio_u$s'] = price_usd

# 2. PER
eps = safe_div(latest.get('NetIncome'), shares) if shares else None
R['PER'] = safe_div(price_ars, eps) if eps else None
R['EPS'] = eps

# 3. Precio Máximo 52 semanas
R['Precio_Max_52s'] = year_high / fx if year_high else None

# 4. Dif. contra Máx 52 sem
R['Dif_Max'] = safe_div(price_ars - year_high, year_high) if year_high else None

# 5. Precio Mínimo 52 semanas
R['Precio_Min_52s'] = year_low / fx if year_low else None

# 6. Dif. contra Mín. 52 sem
R['Dif_Min'] = safe_div(price_ars - year_low, year_low) if year_low else None

# 7. Deuda/EBITDA
total_debt = latest.get('DebtCurrent', 0 or 0) + latest.get('DebtNonCurrent', 0 or 0)
ebitda = latest.get('EBITDA', 0)
R['Deuda_EBITDA'] = safe_div(total_debt, ebitda) if ebitda else None

# 8. Margen Neto
R['Margen_Neto'] = safe_div(latest.get('NetIncome'), latest.get('Revenue'))

# 9. ROE (usando equity promedio si hay años múltiples)
if len(data_by_year) >= 2:
    yrs = list(data_by_year.keys())
    eqs = [data_by_year[y].get('Equity', 0) for y in yrs[-2:]]
    avg_eq = sum(eqs) / len(eqs) if eqs else latest.get('Equity')
    R['ROE'] = safe_div(latest.get('NetIncome'), avg_eq)
else:
    R['ROE'] = safe_div(latest.get('NetIncome'), latest.get('Equity'))

# 10. Crecimiento EPS 5 años
eps_series = []
for yr, d in data_by_year.items():
    ni = d.get('NetIncome')
    if ni is not None and shares:
        eps_series.append((yr, ni / shares))
eps_series.sort()
if len(eps_series) >= 2:
    eps_0 = eps_series[0][1]
    eps_n = eps_series[-1][1]
    n = len(eps_series) - 1
    if eps_0 > 0 and eps_n > 0 and n > 0:
        R['Crec_EPS_5y'] = (eps_n / eps_0) ** (1 / n) - 1
    elif eps_0 < 0 and eps_n > 0:
        R['Crec_EPS_5y'] = None  # Can't calculate from negative base
    else:
        R['Crec_EPS_5y'] = safe_div(eps_n - eps_0, abs(eps_0)) if eps_0 != 0 else None
else:
    R['Crec_EPS_5y'] = None

# 11. FCF / CE (aproximado: NI + DA = FCF / (Equity + Debt))
da = latest.get('DA', 0)
fcf = latest.get('NetIncome', 0) + da
capital_employed = (latest.get('Equity', 0) or 0) + total_debt
R['FCF_CE'] = safe_div(fcf, capital_employed) if capital_employed else None

# 12. Payout (no disponible directamente, dejamos como None)
R['Payout'] = None

# ── Print results ──
print(f"\n  {'='*50}")
print(f"  DASHBOARD GRIM (basado en EEFF al {latest_yr})")
print(f"  {'='*50}")
items = [
    ('Precio en u$s', R.get('Precio_u$s'), 'USD'),
    ('PER', R.get('PER'), 'x'),
    ('Precio Máx 52 sem (USD)', R.get('Precio_Max_52s'), 'USD'),
    ('Dif. contra Máx', R.get('Dif_Max'), '%'),
    ('Precio Mín 52 sem (USD)', R.get('Precio_Min_52s'), 'USD'),
    ('Dif. contra Mín', R.get('Dif_Min'), '%'),
    ('Deuda/EBITDA', R.get('Deuda_EBITDA'), 'x'),
    ('EPS (ARS)', R.get('EPS'), 'ARS'),
    ('Crec. EPS 5y', R.get('Crec_EPS_5y'), '%'),
    ('Margen Neto', R.get('Margen_Neto'), '%'),
    ('ROE', R.get('ROE'), '%'),
    ('FCF/CE', R.get('FCF_CE'), '%'),
    ('Payout', R.get('Payout'), '%'),
]
for name, val, unit in items:
    if val is not None:
        if unit == '%':
            print(f"  {name:<30} = {val:>10.2%}")
        elif unit == 'USD':
            print(f"  {name:<30} = ${val:>10.2f}")
        elif unit == 'ARS':
            print(f"  {name:<30} = ${val:>10.2f}")
        else:
            print(f"  {name:<30} = {val:>10.2f}{unit}")
    else:
        print(f"  {name:<30} = {'N/A':>10}")

# ── FASE 5: Build Excel dashboard ──
print(f"\nFASE 5: Generando dashboard Excel...")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='Calibri', size=11)
title_font = Font(name='Calibri', bold=True, size=14)
sub_font = Font(name='Calibri', bold=True, size=12, color='2F5496')
thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
g_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

ws.merge_cells('A1:O1')
ws['A1'] = f'Dashboard GRIM - Grimoldi SA'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:O2')
ws['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")} | EEFF al {latest_yr} | ARS/USD: {fx}'
ws['A2'].font = Font(name='Calibri', size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

headers = [
    'Especie', 'Listada en', 'Precio u$s', 'PER',
    'Máx 52 sem', 'Dif. Máx', 'Mín 52 sem', 'Dif. Mín',
    'Deuda/EBITDA', 'EPS (ARS)', 'Crec. EPS 5y', 'Margen Neto',
    'ROE', 'FCF/CE', 'Payout'
]

row = 4
for col, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin_border

widths = [8, 10, 12, 8, 14, 10, 14, 10, 12, 12, 14, 12, 10, 10, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def set_cell(r, c, val, fmt=None):
    cell = ws.cell(row=r, column=c)
    if val is not None:
        cell.value = val
        cell.font = data_font
        if fmt:
            cell.number_format = fmt
    else:
        cell.value = 'N/A'
        cell.font = Font(name='Calibri', size=11, color='999999')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

r = 5
set_cell(r, 1, 'GRIM')
set_cell(r, 2, 'BYMA')
set_cell(r, 3, R.get('Precio_u$s'), '$#,##0.00')
set_cell(r, 4, R.get('PER'), '#,##0.00')
set_cell(r, 5, R.get('Precio_Max_52s'), '$#,##0.00')
set_cell(r, 6, R.get('Dif_Max'), '0.00%')
set_cell(r, 7, R.get('Precio_Min_52s'), '$#,##0.00')
set_cell(r, 8, R.get('Dif_Min'), '0.00%')
set_cell(r, 9, R.get('Deuda_EBITDA'), '#,##0.00')
set_cell(r, 10, R.get('EPS'), '#,##0.00')
set_cell(r, 11, R.get('Crec_EPS_5y'), '0.00%')
set_cell(r, 12, R.get('Margen_Neto'), '0.00%')
set_cell(r, 13, R.get('ROE'), '0.00%')
set_cell(r, 14, R.get('FCF_CE'), '0.00%')
set_cell(r, 15, R.get('Payout'))

# Detail section
row = 7
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = 'Detalle Financiero (ARS)'
ws[f'A{row}'].font = sub_font
ws[f'A{row}'].fill = g_fill

r = 8
for col, h in enumerate(['Concepto', 'Último año', 'Año ant.', 'Var. %'], 1):
    c = ws.cell(row=r, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border

fin_items = [
    ('Revenue', 'Ingresos'),
    ('NetIncome', 'Resultado Neto'),
    ('EBITDA', 'EBITDA'),
    ('Equity', 'Patrimonio Neto'),
    ('Assets', 'Activo Total'),
    ('DebtCurrent', 'Deuda CP'),
    ('Cash', 'Caja'),
]
yrs = list(data_by_year.keys())
latest_y = data_by_year[yrs[-1]]
prev_y = data_by_year[yrs[-2]] if len(yrs) >= 2 else {}

for i, (key, label) in enumerate(fin_items):
    rr = r + 1 + i
    ws.cell(row=rr, column=1, value=label).font = data_font
    ws.cell(row=rr, column=1).border = thin_border
    v1 = latest_y.get(key)
    v2 = prev_y.get(key)
    ws.cell(row=rr, column=2, value=v1).font = data_font
    ws.cell(row=rr, column=2).number_format = '#,##0'
    ws.cell(row=rr, column=2).border = thin_border
    if v2 is not None:
        ws.cell(row=rr, column=3, value=v2).font = data_font
        ws.cell(row=rr, column=3).number_format = '#,##0'
    else:
        ws.cell(row=rr, column=3, value='N/A').font = Font(color='999999')
    ws.cell(row=rr, column=3).border = thin_border
    if v1 is not None and v2 is not None and v2 != 0:
        ws.cell(row=rr, column=4, value=(v1-v2)/abs(v2)).font = data_font
        ws.cell(row=rr, column=4).number_format = '0.00%'
    else:
        ws.cell(row=rr, column=4, value='N/A').font = Font(color='999999')
    ws.cell(row=rr, column=4).border = thin_border

# Multi-year history section
row = 18
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = 'Histórico 5 años'
ws[f'A{row}'].font = sub_font
ws[f'A{row}'].fill = g_fill

r = 19
hist_headers = ['Año'] + [f for f, _ in fin_items[:4]]
for col, h in enumerate(hist_headers, 1):
    c = ws.cell(row=r, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border

for i, yr in enumerate(reversed(yrs)):
    rr = r + 1 + i
    d = data_by_year[yr]
    vals = [yr] + [d.get(k, 0) for k, _ in fin_items[:4]]
    for j, v in enumerate(vals):
        cell = ws.cell(row=rr, column=j+1)
        if j == 0:
            cell.value = v
        else:
            cell.value = v
            cell.number_format = '#,##0'
        cell.font = data_font
        cell.border = thin_border

output_path = BASE / 'Dashboard_GRIM_v2.xlsx'
wb.save(output_path)

print(f"\n{'='*70}")
print(f"DASHBOARD COMPLETO: {output_path}")
print(f"{'='*70}")
print(f"\nResumen GRIM (Grimoldi SA):")
print(f"  Precio: USD${R.get('Precio_u$s',0):.2f}")
print(f"  PER: {R.get('PER', 'N/A')}")
print(f"  Margen Neto: {R.get('Margen_Neto', 'N/A'):.2%}" if R.get('Margen_Neto') else "  Margen Neto: N/A")
print(f"  ROE: {R.get('ROE', 'N/A'):.2%}" if R.get('ROE') else "  ROE: N/A")
print(f"  Deuda/EBITDA: {R.get('Deuda_EBITDA', 'N/A'):.2f}x" if R.get('Deuda_EBITDA') else "  Deuda/EBITDA: N/A")
print(f"  EPS: ARS${R.get('EPS', 0):.2f}")
print(f"  Crec. EPS 5y: {R.get('Crec_EPS_5y', 'N/A'):.2%}" if R.get('Crec_EPS_5y') else "  Crec. EPS 5y: N/A")
