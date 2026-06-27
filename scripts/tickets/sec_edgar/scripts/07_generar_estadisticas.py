"""
PASO 7 — Generar la solapa "estadisticas_outliers": distribucion de la
divergencia EDGAR-vs-Investing por metrica (media, desv. std, P50/P90/P95/P99)
y un punch-list de outliers ticker-por-ticker con ingenieria inversa para
priorizar que vale la pena investigar/arreglar.

IMPORTANTE: la divergencia se recalcula ACA, desde cero, leyendo Investing
crudo (resultados!B:G) + datos/ratios_tickets.json (EDGAR). NO se usa
resultados!Q:V (Comparacion %) como fuente: esa columna esta desincronizada
con el estado actual del calculo (ej. CAT tiene valores EDGAR en I-O pero
Q-V vacio; GS mostraba Margen "igual" a Investing cuando en realidad EDGAR
no tiene ese dato). Recalcular desde la fuente evita heredar ese problema.

Ingenieria inversa (lo que pidio el usuario -- "que datos podemos replicar
desde SEC EDGAR a Investing"): para cada outlier, se asume que el numero de
Investing es correcto y se despeja el insumo que haria falta para que EDGAR
coincida (ej. en PER: que EPS necesitaria Investing, y que Net Income
implica eso dado el share count de EDGAR). Comparar ese insumo implicito
contra el insumo real de EDGAR es lo que dice POR DONDE esta la diferencia.

Input : informe_seguimiento_detallado.xlsx (hoja resultados, columnas B-G)
        datos/ratios_tickets.json
Output: nueva hoja "estadisticas_outliers" en el mismo .xlsx
"""

from __future__ import annotations

import json
import math
import shutil
from pathlib import Path

import openpyxl

ROOT = next(p for p in Path(__file__).resolve().parents if (p / 'data').is_dir())
XLSX_PATH = ROOT / "informe_seguimiento_detallado.xlsx"
TMP_PATH = ROOT / "informe_seguimiento_detallado.tmp.xlsx"
RATIOS_JSON = Path(__file__).resolve().parent.parent / "datos" / "ratios_tickets.json"
SHEET_NAME = "estadisticas_outliers"

# nombre -> (col Investing en 'resultados' [1-idx], campo EDGAR en ratios_tickets.json, es_porcentaje)
METRICAS = {
    "PER":      (2, "per_ttm", False),
    "EPS":      (3, "eps_annual", False),
    "CrecEPS":  (4, "cagr_eps_5y", True),
    "Margen":   (5, "margen_neto_ttm", True),
    "ROE":      (6, "roe_ttm", True),
    "Payout":   (7, "payout_ttm", True),
}


# ============================================================
# ESTADISTICA (sin numpy/scipy, no son dependencias del proyecto)
# ============================================================

def percentil(valores_ordenados: list[float], p: float) -> float:
    """Percentil por interpolacion lineal (metodo comun, igual a numpy 'linear')."""
    n = len(valores_ordenados)
    if n == 0:
        return float("nan")
    if n == 1:
        return valores_ordenados[0]
    idx = (p / 100) * (n - 1)
    lo = math.floor(idx)
    hi = math.ceil(idx)
    if lo == hi:
        return valores_ordenados[lo]
    frac = idx - lo
    return valores_ordenados[lo] + (valores_ordenados[hi] - valores_ordenados[lo]) * frac


def media(vals: list[float]) -> float:
    return sum(vals) / len(vals) if vals else float("nan")


def desv_std(vals: list[float]) -> float:
    if len(vals) < 2:
        return float("nan")
    m = media(vals)
    var = sum((v - m) ** 2 for v in vals) / (len(vals) - 1)
    return math.sqrt(var)


def resumen_metrica(divergencias: list[float]) -> dict:
    """divergencias: % con signo (EDGAR-Investing)/Investing*100."""
    n = len(divergencias)
    if n == 0:
        return {"n": 0}
    abs_ord = sorted(abs(d) for d in divergencias)
    p95 = percentil(abs_ord, 95)
    p99 = percentil(abs_ord, 99)
    return {
        "n": n,
        "media_con_signo": media(divergencias),
        "desv_std": desv_std(divergencias),
        "min_abs": abs_ord[0],
        "p50_abs": percentil(abs_ord, 50),
        "p90_abs": percentil(abs_ord, 90),
        "p95_abs": p95,
        "p99_abs": p99,
        "max_abs": abs_ord[-1],
        "pct_le_5": sum(1 for v in abs_ord if v <= 5) / n * 100,
        "pct_le_10": sum(1 for v in abs_ord if v <= 10) / n * 100,
        "pct_le_20": sum(1 for v in abs_ord if v <= 20) / n * 100,
        "umbral_outlier_p95": p95,
        "umbral_outlier_p99": p99,
    }


# ============================================================
# CARGA DE DATOS
# ============================================================

def cargar_investing() -> dict[str, dict]:
    """{ticker: {PER, EPS, CrecEPS, Margen, ROE, Payout}} desde resultados!B:G."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["resultados"]
    out = {}
    for r in range(3, ws.max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        if not ticker:
            continue
        fila = {}
        for nombre, (col_inv, _, _) in METRICAS.items():
            fila[nombre] = ws.cell(row=r, column=col_inv).value
        out[ticker] = fila
    return out


def a_numero(v) -> float | None:
    """Coerciona a float; descarta texto no numerico (ej 'JAMAICA O SIDNY')
    o errores de formula viejos tipo '#VALUE!' que puedan quedar en celdas."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


def cargar_edgar() -> dict[str, dict]:
    with open(RATIOS_JSON, encoding="utf-8") as f:
        filas = json.load(f)["filas"]
    return {f["ticker"]: f for f in filas}


# ============================================================
# INGENIERIA INVERSA -- por outlier, que insumo EDGAR explicaria la brecha
# ============================================================

def reverse_engineer(metrica: str, ticker: str, inv_val: float, edgar_val: float,
                      edgar_row: dict) -> tuple[str, float | None, float | None, float | None]:
    """
    Devuelve (descripcion_insumo, insumo_implicito_desde_investing,
    insumo_real_edgar, diff_pct). diff_pct = None si no se puede reconstruir.
    """
    precio = edgar_row.get("precio_usd")
    shares = edgar_row.get("_diluted_shares")
    ni_edgar = edgar_row.get("_netincome_ttm")
    rev_edgar = edgar_row.get("_revenue_ttm")
    eq_prom = edgar_row.get("_equity_prom")

    def diff(implicito, real):
        if implicito is None or real is None or real == 0:
            return None
        return (implicito - real) / real * 100

    if metrica == "PER":
        if precio and inv_val and shares:
            eps_implicito = precio / inv_val
            ni_implicito = eps_implicito * shares
            return ("NetIncome_TTM implicito (via EPS=Precio/PER_inv * Shares_EDGAR)",
                    ni_implicito, ni_edgar, diff(ni_implicito, ni_edgar))
        return ("NetIncome_TTM implicito", None, ni_edgar, None)

    if metrica == "EPS":
        if shares:
            ni_implicito = inv_val * shares
            return ("NetIncome_TTM implicito (via EPS_inv * Shares_EDGAR)",
                    ni_implicito, ni_edgar, diff(ni_implicito, ni_edgar))
        return ("NetIncome_TTM implicito", None, ni_edgar, None)

    if metrica == "Margen":
        if rev_edgar:
            ni_implicito = inv_val * rev_edgar
            return ("NetIncome_TTM implicito (via Margen_inv * Revenue_TTM_EDGAR)",
                    ni_implicito, ni_edgar, diff(ni_implicito, ni_edgar))
        return ("NetIncome_TTM implicito", None, ni_edgar, None)

    if metrica == "ROE":
        if ni_edgar:
            eq_implicito = ni_edgar / inv_val if inv_val else None
            return ("Equity_Promedio implicito (via NetIncome_EDGAR / ROE_inv)",
                    eq_implicito, eq_prom, diff(eq_implicito, eq_prom))
        return ("Equity_Promedio implicito", None, eq_prom, None)

    if metrica == "Payout":
        if ni_edgar:
            div_implicito = inv_val * ni_edgar
            div_edgar = edgar_row.get("_div_ttm")
            return ("Dividendos_TTM implicito (via Payout_inv * NetIncome_EDGAR)",
                    div_implicito, div_edgar, diff(div_implicito, div_edgar))
        return ("Dividendos_TTM implicito", None, edgar_row.get("_div_ttm"), None)

    # CrecEPS: CAGR de 2 puntos, no se puede despejar un insumo unico con una
    # sola ecuacion -- ver docs/screener/DIAGNOSTICO_INVESTING_vs_EDGAR.md #3.C
    return ("(no reconstruible: CrecEPS depende de 2 puntos temporales, "
            "metodologia de Investing distinta -- ver DIAGNOSTICO_INVESTING_vs_EDGAR.md)",
            None, None, None)


def hipotesis_texto(metrica: str, diff_pct: float | None, descripcion: str) -> str:
    if diff_pct is None:
        return "No se pudo reconstruir un insumo implicito (falta dato EDGAR base)."
    if abs(diff_pct) <= 10:
        return (f"{descripcion}: coincide razonablemente con EDGAR ({diff_pct:+.1f}%) -- "
                f"la brecha del ratio probablemente es de ventana temporal (TTM vs FY) "
                f"o redondeo, no un insumo distinto.")
    return (f"{descripcion} difiere {diff_pct:+.1f}% del valor real de EDGAR -- "
            f"esa es la pieza concreta a revisar (¿item no recurrente? ¿tag XBRL "
            f"equivocado? ¿periodo distinto?).")


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    print("=" * 60)
    print("  PASO 7 -- Generar solapa estadisticas_outliers")
    print("=" * 60)

    investing = cargar_investing()
    edgar = cargar_edgar()
    print(f"\n  Tickers en resultados (Investing): {len(investing)}")
    print(f"  Tickers en ratios_tickets.json (EDGAR): {len(edgar)}")

    # --- recalcular divergencias por metrica ---
    divergencias_por_metrica: dict[str, list[tuple[str, float, float, float]]] = {m: [] for m in METRICAS}

    for ticker, inv_fila in investing.items():
        edgar_row = edgar.get(ticker)
        if not edgar_row or edgar_row.get("_error"):
            continue
        for metrica, (_, campo_edgar, es_pct) in METRICAS.items():
            inv_val = a_numero(inv_fila.get(metrica))
            edgar_val = a_numero(edgar_row.get(campo_edgar))
            if inv_val is None or edgar_val is None or inv_val == 0:
                continue
            diff_pct = (edgar_val - inv_val) / inv_val * 100
            divergencias_por_metrica[metrica].append((ticker, inv_val, edgar_val, diff_pct))

    resumenes = {}
    for metrica, filas in divergencias_por_metrica.items():
        resumenes[metrica] = resumen_metrica([f[3] for f in filas])
        print(f"  {metrica:10s} n={resumenes[metrica]['n']:3d}")

    # --- armar tabla de outliers (>P95 de |divergencia| para esa metrica) ---
    outliers = []
    for metrica, filas in divergencias_por_metrica.items():
        resumen = resumenes[metrica]
        if resumen["n"] == 0:
            continue
        umbral95 = resumen["umbral_outlier_p95"]
        umbral99 = resumen["umbral_outlier_p99"]
        for ticker, inv_val, edgar_val, diff_pct in filas:
            if abs(diff_pct) < umbral95:
                continue
            edgar_row = edgar[ticker]
            desc, implicito, real_edgar, diff_implicito = reverse_engineer(
                metrica, ticker, inv_val, edgar_val, edgar_row
            )
            outliers.append({
                "ticker": ticker,
                "metrica": metrica,
                "investing": inv_val,
                "edgar": edgar_val,
                "diff_pct": diff_pct,
                "es_extremo_p99": abs(diff_pct) >= umbral99,
                "insumo_a_revisar": desc,
                "insumo_implicito": implicito,
                "insumo_real_edgar": real_edgar,
                "diff_insumo_pct": diff_implicito,
                "hipotesis": hipotesis_texto(metrica, diff_implicito, desc),
            })

    outliers.sort(key=lambda o: abs(o["diff_pct"]), reverse=True)
    print(f"\n  Outliers detectados (>P95 de su metrica): {len(outliers)}")

    # ============================================================
    # ESCRITURA EN EL XLSX
    # ============================================================
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    sheets_originales = list(wb.sheetnames)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
        print(f"  (ya existia '{SHEET_NAME}', se recrea)")

    ws = wb.create_sheet(SHEET_NAME)

    # --- Bloque 1: resumen estadistico por metrica ---
    ws.cell(row=1, column=1, value="RESUMEN ESTADISTICO -- divergencia % (EDGAR-Investing)/Investing*100, recalculada desde cero")
    headers1 = ["Metrica", "n", "Media (con signo)", "Desv.Std", "Min |dif|", "P50 |dif|",
                "P90 |dif|", "P95 |dif|", "P99 |dif|", "Max |dif|",
                "% <=5%", "% <=10%", "% <=20%"]
    for c, h in enumerate(headers1, start=1):
        ws.cell(row=2, column=c, value=h)

    r = 3
    for metrica in METRICAS:
        s = resumenes[metrica]
        if s["n"] == 0:
            ws.cell(row=r, column=1, value=metrica)
            ws.cell(row=r, column=2, value=0)
            r += 1
            continue
        valores = [metrica, s["n"], s["media_con_signo"], s["desv_std"], s["min_abs"],
                   s["p50_abs"], s["p90_abs"], s["p95_abs"], s["p99_abs"], s["max_abs"],
                   s["pct_le_5"], s["pct_le_10"], s["pct_le_20"]]
        for c, v in enumerate(valores, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    r += 2
    fila_inicio_outliers = r
    ws.cell(row=r, column=1,
            value="OUTLIERS (>P95 de su metrica) -- punch-list ordenado por |divergencia|, con ingenieria inversa")
    r += 1
    headers2 = ["Ticker", "Metrica", "Investing", "EDGAR", "Diff %", "Extremo (>P99)",
                "Insumo a revisar", "Insumo implicito (desde Investing)",
                "Insumo real EDGAR", "Diff insumo %", "Hipotesis"]
    for c, h in enumerate(headers2, start=1):
        ws.cell(row=r, column=c, value=h)
    r += 1

    for o in outliers:
        valores = [o["ticker"], o["metrica"], o["investing"], o["edgar"], o["diff_pct"],
                   "SI" if o["es_extremo_p99"] else "",
                   o["insumo_a_revisar"], o["insumo_implicito"], o["insumo_real_edgar"],
                   o["diff_insumo_pct"], o["hipotesis"]]
        for c, v in enumerate(valores, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    for c in range(1, 14):
        letra = ws.cell(row=2, column=c).column_letter
        ws.column_dimensions[letra].width = 18
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["K"].width = 60

    print(f"  Tabla de outliers escrita desde la fila {fila_inicio_outliers}")

    wb.save(TMP_PATH)

    # --- validacion antes de reemplazar ---
    wb_check = openpyxl.load_workbook(TMP_PATH, data_only=False)
    if set(wb_check.sheetnames) != set(sheets_originales) | {SHEET_NAME}:
        TMP_PATH.unlink(missing_ok=True)
        raise SystemExit("ABORTADO: las hojas no coinciden con lo esperado. Archivo real intacto.")
    if wb_check["resultados"]["I3"].value != 18.36:
        TMP_PATH.unlink(missing_ok=True)
        raise SystemExit("ABORTADO: 'resultados' cambio inesperadamente. Archivo real intacto.")
    if "calculos_edgar" not in wb_check.sheetnames:
        TMP_PATH.unlink(missing_ok=True)
        raise SystemExit("ABORTADO: 'calculos_edgar' desaparecio. Archivo real intacto.")
    print("  Validacion OK: el resto de las hojas quedo intacto "
          "(no se valida el contenido de 'calculos_edgar' celda por celda -- "
          "este script no la toca, pudiste haberla editado a mano).")

    shutil.move(str(TMP_PATH), str(XLSX_PATH))
    print(f"\n  OK {XLSX_PATH}  (solapa '{SHEET_NAME}' agregada)")
    print("\nPaso 7 completo.")


if __name__ == "__main__":
    main()
