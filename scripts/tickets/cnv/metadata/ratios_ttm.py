# -*- coding: utf-8 -*-
"""ratios_ttm: recalcula TODOS los ratios de flujo de las byma_only con el
numerador en TTM (des-acumulado) y el denominador de balance (snapshot del ultimo
periodo). Escribe la tabla NUEVA `ratios_ttm` (no toca `screener`) y genera un
Excel de comparacion lado a lado (screener vs TTM) + los inputs crudos para verificar.

    python scripts/tickets/cnv/metadata/ratios_ttm.py

Requiere fiscal_calendar (build_fiscal_calendar.py). No sobrescribe nada existente.
"""
from __future__ import annotations
import os as _os
import os, sqlite3, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", ".."))
DB = os.path.join(ROOT, "data", _os.environ.get("SCREENER_DB", "screener.db"))
XLSX = os.path.join(ROOT, "docs", "proyecto", "Ratios_TTM_vs_Screener.xlsx")

FLOW = ["NetIncome", "Revenue", "OperatingIncome", "GrossProfit", "EBITDA", "CF_Operativo", "CF_Inversion"]
BAL = ["Equity", "Assets", "DebtCurrent", "DebtNonCurrent", "Cash"]


def mb(a, b): return (int(b[:4]) - int(a[:4])) * 12 + (int(b[5:7]) - int(a[5:7]))
def qnum(m, fy): return 4 - ((fy - m) % 12) // 3


def series(cur, cuit, c):
    d = {}
    for pe, v in cur.execute("select period_end,valor from cnv_estados_v2 where cuit=? and concepto=? and valor is not null order by period_end", (cuit, c)):
        d[pe] = v
    return sorted(d.items())


def ttm(cur, cuit, c, fy):
    """suma de los 4 ultimos trimestres standalone (des-acumulado)."""
    ser = series(cur, cuit, c)
    out, prev = [], None
    sa = []
    for pe, cum in ser:
        q = qnum(int(pe[5:7]), fy)
        v = cum if q == 1 else (cum - prev[1] if (prev and mb(prev[0], pe) == 3) else None)
        sa.append((pe, v)); prev = (pe, cum)
    sa = [(pe, v) for pe, v in sa if v is not None]
    if len(sa) < 4: return None
    w = sa[-4:]
    if any(mb(w[i][0], w[i + 1][0]) != 3 for i in range(3)): return None
    return sum(v for _, v in w)


def snap(cur, cuit, c):
    r = cur.execute("select valor from cnv_estados_v2 where cuit=? and concepto=? order by period_end desc limit 1", (cuit, c)).fetchone()
    return r[0] if r else None


def compute():
    con = sqlite3.connect(DB); cur = con.cursor()
    fc = {cu: fy for cu, fy in cur.execute("select cuit,fy_end_month from fiscal_calendar")}
    uni = cur.execute("select cuit,ticker,ultimo_periodo,MarketCapUSD,ROE,ROA,MargenNeto,margen_operativo,DeudaEBITDA,ev_ebitda,PER from screener where grupo='byma_only' order by ticker").fetchall()
    rows = []
    for cuit, tk, up, mcap, s_roe, s_roa, s_mgn, s_mgo, s_de, s_ev, s_per in uni:
        fy = fc.get(cuit)
        interino = int(bool(up and fy and int(up[5:7]) != fy))
        f = {c: (ttm(cur, cuit, c, fy) if fy else None) for c in FLOW}
        b = {c: snap(cur, cuit, c) for c in BAL}
        ni, rev, oi, gp, eb, cfo, cfi = (f[c] for c in FLOW)
        eq, ast, dc, dnc, cash = (b[c] for c in BAL)
        mcap = mcap if isinstance(mcap, (int, float)) else None
        deuda = (dc or 0) + (dnc or 0)
        def div(a, x, pos=False):
            return (a / x) if (isinstance(a, (int, float)) and isinstance(x, (int, float)) and x and (x > 0 if pos else True)) else None
        roe = div(ni, eq, pos=True)
        roa = div(ni, ast, pos=True)
        mgn = div(ni, rev, pos=True)
        mgo = div(oi, rev, pos=True)
        mgb = div(gp, rev, pos=True)
        de = div(deuda, eb, pos=True) if (isinstance(eb, (int, float)) and eb > 0) else None
        ev = ((mcap or 0) + deuda - (cash or 0))
        eve = (ev / eb) if (isinstance(eb, (int, float)) and eb > 0 and mcap) else None
        ce = deuda + (eq or 0)
        fcf = (cfo or 0) + (cfi or 0)
        fcfce = (fcf / ce) if ce else None
        per = (mcap / ni) if (mcap and isinstance(ni, (int, float)) and ni > 0) else None
        # gates de sanidad: fuera de rango = denominador roto (ej. patrimonio ~0) -> None
        def cap(x, lo, hi): return x if (isinstance(x, (int, float)) and lo <= x <= hi) else None
        roe = cap(roe, -10, 10); roa = cap(roa, -10, 10)          # |ROE/ROA| <= 1000%
        mgn = cap(mgn, -10, 10); mgo = cap(mgo, -10, 10); mgb = cap(mgb, -10, 10)
        de = cap(de, 0, 100); eve = cap(eve, -100, 100); fcfce = cap(fcfce, -10, 10)
        per = cap(per, 1, 100)
        rows.append(dict(cuit=cuit, ticker=tk, interino=interino,
                         roe=roe, roa=roa, mgn=mgn, mgo=mgo, mgb=mgb, de=de, eve=eve, fcfce=fcfce, per=per,
                         ni=ni, rev=rev, oi=oi, eb=eb, cfo=cfo, cfi=cfi, eq=eq, ast=ast, deuda=deuda, cash=cash, mcap=mcap,
                         s_roe=s_roe, s_roa=s_roa, s_mgn=s_mgn, s_mgo=s_mgo, s_de=s_de, s_ev=s_ev, s_per=s_per))
    return con, cur, rows


def persist(con, cur, rows):
    cur.executescript("""
        DROP TABLE IF EXISTS ratios_ttm;
        CREATE TABLE ratios_ttm(
            cuit TEXT PRIMARY KEY, ticker TEXT, interino INTEGER,
            roe REAL, roa REAL, margen_neto REAL, margen_op REAL, margen_bruto REAL,
            deuda_ebitda REAL, ev_ebitda REAL, fcf_ce REAL, per REAL,
            ni_ttm REAL, rev_ttm REAL, oi_ttm REAL, ebitda_ttm REAL, cfo_ttm REAL, cfi_ttm REAL,
            equity REAL, assets REAL, deuda REAL, cash REAL, market_cap REAL, built_at TEXT);
    """)
    now = dt.datetime.now().isoformat(timespec="seconds")
    cur.executemany("INSERT INTO ratios_ttm VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(r["cuit"], r["ticker"], r["interino"], r["roe"], r["roa"], r["mgn"], r["mgo"], r["mgb"],
          r["de"], r["eve"], r["fcfce"], r["per"], r["ni"], r["rev"], r["oi"], r["eb"], r["cfo"], r["cfi"],
          r["eq"], r["ast"], r["deuda"], r["cash"], r["mcap"], now) for r in rows])
    con.commit()


# ---------------- Excel ----------------
NAVY="1A2744"; WHITE="FFFFFF"; GRID="D6E2F2"; GREEN_BG="E7F7EC"; GREEN_T="1F6B35"; RED="C0392B"; INK="1A2744"; GRAYL="6B7589"; LB="DCEAFF"
def Ft(sz=10,b=False,color="3A4458"): return Font(name="Calibri",size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
S=Side(style="thin",color=GRID); BORDER=Border(left=S,right=S,top=S,bottom=S)
CEN=Alignment(horizontal="center",vertical="center"); LEFT=Alignment(horizontal="left",vertical="center")


def _num(ws,r,c,v,fmt=None,font=None):
    cell=ws.cell(r,c, round(v,3) if isinstance(v,(int,float)) else None)
    if fmt and isinstance(v,(int,float)): cell.number_format=fmt
    cell.font=font or Ft(9); cell.border=BORDER; cell.alignment=CEN
    return cell


def excel(rows):
    wb=Workbook(); ws=wb.active; ws.title="Comparacion"
    ws.column_dimensions["A"].width=9; ws.column_dimensions["B"].width=9
    PAIRS=[("ROE","roe","s_roe","0.0%"),("ROA","roa","s_roa","0.0%"),("Mg Neto","mgn","s_mgn","0.0%"),
           ("Mg Op.","mgo","s_mgo","0.0%"),("Deuda/EBITDA","de","s_de","0.0"),("EV/EBITDA","eve","s_ev","0.0"),
           ("PER","per","s_per","0.0")]
    ws.merge_cells("A1:B1"); ws["A1"]="Screener (viejo) vs TTM (corregido)"; ws["A1"].font=Ft(11,True,NAVY)
    # fila 2: grupos
    ws.cell(3,1,"Ticker").fill=fill(NAVY); ws.cell(3,1).font=Ft(9.5,True,WHITE); ws.cell(3,1).alignment=CEN; ws.cell(3,1).border=BORDER
    ws.cell(3,2,"Interino").fill=fill(NAVY); ws.cell(3,2).font=Ft(9,True,WHITE); ws.cell(3,2).alignment=CEN; ws.cell(3,2).border=BORDER
    col=3
    for nm,_,_,_ in PAIRS:
        ws.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+1)
        h=ws.cell(2,col,nm); h.fill=fill(NAVY); h.font=Ft(9.5,True,WHITE); h.alignment=CEN; h.border=BORDER
        ws.cell(3,col,"scr").fill=fill("4A5573"); ws.cell(3,col).font=Ft(8,True,WHITE); ws.cell(3,col).alignment=CEN; ws.cell(3,col).border=BORDER
        ws.cell(3,col+1,"TTM").fill=fill(GREEN_T); ws.cell(3,col+1).font=Ft(8,True,WHITE); ws.cell(3,col+1).alignment=CEN; ws.cell(3,col+1).border=BORDER
        ws.column_dimensions[get_column_letter(col)].width=8; ws.column_dimensions[get_column_letter(col+1)].width=8
        col+=2
    r=4
    for row in rows:
        ws.cell(r,1,row["ticker"]).font=Ft(9.5,True,INK); ws.cell(r,1).border=BORDER; ws.cell(r,1).alignment=LEFT
        it=ws.cell(r,2,"SÍ" if row["interino"] else "no"); it.font=Ft(8.5,True,RED if row["interino"] else GREEN_T); it.border=BORDER; it.alignment=CEN
        col=3
        for _,tk,sk,fmt in PAIRS:
            _num(ws,r,col,row[sk],fmt,Ft(9,color=GRAYL))
            _num(ws,r,col+1,row[tk],fmt,Ft(9,True,GREEN_T)).fill=fill(GREEN_BG)
            col+=2
        r+=1
    ws.freeze_panes="C4"

    # inputs para verificar
    iw=wb.create_sheet("Inputs_TTM")
    heads=["Ticker","NI TTM","Rev TTM","OI TTM","EBITDA TTM","CFO TTM","CFI TTM","Equity","Assets","Deuda","Cash","MarketCap"]
    keys=["ticker","ni","rev","oi","eb","cfo","cfi","eq","ast","deuda","cash","mcap"]
    for i,w in enumerate([9]+[13]*11,1): iw.column_dimensions[get_column_letter(i)].width=w
    for i,h in enumerate(heads,1):
        c=iw.cell(1,i,h); c.fill=fill(NAVY); c.font=Ft(9,True,WHITE); c.alignment=CEN; c.border=BORDER
    iw.cell(2,1,"(sumá los 4 trim. standalone de Tags_BYMA y tiene que dar el TTM de acá)").font=Ft(8,False,GRAYL)
    r=3
    for row in rows:
        for i,k in enumerate(keys,1):
            v=row[k]; cell=iw.cell(r,i, round(v,0) if isinstance(v,(int,float)) else None)
            cell.font=Ft(9,True,INK) if i==1 else Ft(9); cell.border=BORDER
            cell.alignment=LEFT if i==1 else Alignment(horizontal="right")
        r+=1
    iw.freeze_panes="B2"
    wb.save(XLSX)


def main():
    con, cur, rows = compute()
    persist(con, cur, rows)
    excel(rows)
    # reporte: cuanto cambio en las interinas
    print(f"ratios_ttm: {len(rows)} byma_only  |  Excel: {XLSX}")
    interinas=[r for r in rows if r["interino"]]
    print(f"interinas (ratios que estaban mal): {len(interinas)}")
    print("\nejemplos de correccion (ROE screener -> TTM):")
    for r in sorted(interinas, key=lambda x:-(x['roe'] or -9))[:8]:
        def p(x): return f'{x*100:.1f}%' if isinstance(x,(int,float)) else '—'
        print(f"   {r['ticker']:>6}  scr={p(r['s_roe']):>7}  ->  TTM={p(r['roe']):>7}")
    con.close()


if __name__ == "__main__":
    main()
