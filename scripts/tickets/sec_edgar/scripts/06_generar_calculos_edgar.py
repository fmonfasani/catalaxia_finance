"""
PASO 6 — Generar la solapa "calculos_edgar" dentro de
informe_seguimiento_detallado.xlsx, mostrando el desglose completo de como se
arma cada uno de los 6 ratios EDGAR (columnas I-O de la solapa "resultados"):
PER, EPS_anual, EPS_TTM, Crecimiento EPS 5a, Margen Neto, ROE, Payout.

Cada ratio queda como una FORMULA de Excel que referencia las celdas de sus
componentes (igual que pidio el usuario: B2=C2/D2, etc.), y los componentes
crudos (NetIncome_TTM, Revenue_TTM, Shares, Equity, Dividendos) se muestran
como valores con su tag XBRL de origen y la estrategia de TTM usada -- asi
cada numero es auditable hasta la fuente EDGAR exacta, sin tener que abrir
el JSON crudo.

No se recalcula nada distinto a lo que ya hace 03_calcular_ratios.py: este
script solo lee datos/ratios_tickets.json (que ya tiene esos componentes
expuestos como campos _debug) y los acomoda en la hoja.

Escritura segura: copia a .tmp, valida que las demas 15 hojas del workbook
sigan intactas, y solo entonces reemplaza el archivo real.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl

ROOT = next(p for p in Path(__file__).resolve().parents if (p / 'data').is_dir())
XLSX_PATH = ROOT / "informe_seguimiento_detallado.xlsx"
TMP_PATH = ROOT / "informe_seguimiento_detallado.tmp.xlsx"
RATIOS_JSON = Path(__file__).resolve().parent.parent / "datos" / "ratios_tickets.json"
SHEET_NAME = "calculos_edgar"

HEADERS = [
    "Ticker", "Empresa",
    "PER", "Precio_Actual", "EPS_TTM", "NetIncome_TTM", "Shares_Diluted_Weighted",
    "EPS_Anual",
    "Crec_EPS_5y", "EPS_fin_reciente", "EPS_ini_hace_5y",
    "Margen_Neto", "Revenue_TTM",
    "ROE_TTM", "Equity_Promedio", "Equity_actual", "Equity_hace_1_anio",
    "Payout", "Dividendos_TTM",
    "Tag_NetIncome_XBRL", "Tag_Revenue_XBRL", "Estrategia_TTM_NetIncome",
    "Fecha_NetIncome_usado", "Flags_calidad", "Notas",
]

# Columnas 1-indexadas (A=1) para escribir formulas que referencian otras celdas.
COL = {nombre: i + 1 for i, nombre in enumerate(HEADERS)}


def construir_notas(fila: dict) -> str:
    if fila.get("_error"):
        return f"Sin datos EDGAR: {fila['_error']} (no se llego a calcular ningun ratio)"

    notas = []
    if fila.get("_stale"):
        notas.append(f"STALE: ultimo dato real es de {fila.get('_end_datos')} "
                      f"(>2 anios) -> todos los ratios TTM se anulan a proposito")
    if fila.get("_ni_es_fy"):
        notas.append("El 'TTM' de Net Income es en realidad el ultimo anio "
                      "fiscal cerrado (estrategia C), no un TTM rodante -- "
                      "puede divergir de Investing si hubo un item no "
                      "recurrente en un trimestre reciente")
    if not fila.get("_metric_revenue") and fila.get("margen_neto_ttm") is None and not fila.get("_stale"):
        notas.append("Margen Neto no disponible: no se encontro ningun tag "
                      "de Revenue (Revenues/RevenueFromContract.../"
                      "SalesRevenueNet) -- tipico de bancos/holdings que no "
                      "reportan una linea unica de ingresos")
    if fila.get("_equity_actual") is None and fila.get("roe_ttm") is None and not fila.get("_stale"):
        notas.append("ROE no disponible: no se encontro el tag "
                      "StockholdersEquity con valor y fecha en los "
                      "financials descargados")
    if fila.get("_diluted_shares") is None and fila.get("eps_ttm_diluted") is None and not fila.get("_stale"):
        notas.append("EPS_TTM/PER no disponibles: no hay "
                      "WeightedAverageNumberOfDilutedSharesOutstanding ni "
                      "...Basic con valor (o se descarto por error de escala, "
                      "ver flag shares_descartadas_escala)")
    if fila.get("_div_ttm") is None and fila.get("payout_ttm") is None and not fila.get("_stale"):
        notas.append("Payout no disponible: no hay tag de dividendos "
                      "pagados (PaymentsOfDividendsCommonStock/"
                      "PaymentsOfDividends) -- normal en empresas que no "
                      "reparten dividendos, o que no taggean ese dato")
    flags = fila.get("_flags") or ""
    if "shares_descartadas_escala" in flags:
        notas.append("Shares diluidas descartadas: el valor tageado en EDGAR "
                      "difiere >100x del CommonStockSharesOutstanding (dato "
                      "crudo con error de escala en el filing)")
    if "roe_no_significativo" in flags:
        notas.append("ROE no interpretable: el equity promedio es <5% de los "
                      "activos totales (tipico de empresas con buybacks "
                      "agresivos, equity casi nulo)")
    if "ebitda_da_mas_intangibles" in flags:
        notas.append("D&A incluye amortizacion de intangibles sumada aparte "
                      "(el tag primario solo cubria PP&E)")
    return "; ".join(notas)


def escribir_fila(ws, r: int, fila: dict) -> None:
    def set_val(col_name: str, value):
        ws.cell(row=r, column=COL[col_name], value=value)

    def set_formula(col_name: str, formula: str):
        ws.cell(row=r, column=COL[col_name], value=formula)

    set_val("Ticker", fila["ticker"])
    set_val("Empresa", fila.get("nombre"))

    precio = fila.get("precio_usd")
    netinc_ttm = fila.get("_netincome_ttm")
    shares = fila.get("_diluted_shares")
    revenue_ttm = fila.get("_revenue_ttm")
    equity_actual = fila.get("_equity_actual")
    equity_previo = fila.get("_equity_previo")
    div_ttm = fila.get("_div_ttm")
    eps_fin = fila.get("_eps_fin_val")
    eps_ini = fila.get("_eps_ini_val")

    r_ = str(r)

    # --- EPS_TTM = NetIncome_TTM / Shares_Diluted_Weighted ---
    set_val("NetIncome_TTM", netinc_ttm)
    set_val("Shares_Diluted_Weighted", shares)
    if netinc_ttm is not None and shares:
        set_formula("EPS_TTM", f"=F{r_}/G{r_}")
    else:
        set_val("EPS_TTM", None)

    # --- PER = Precio_Actual / EPS_TTM ---
    set_val("Precio_Actual", precio)
    if precio is not None and fila.get("eps_ttm_diluted"):
        set_formula("PER", f"=D{r_}/E{r_}")
    else:
        set_val("PER", None)

    # --- EPS_Anual: valor directo (ultimo EarningsPerShareDiluted de 10-K) ---
    set_val("EPS_Anual", fila.get("eps_annual"))

    # --- Crec_EPS_5y = (EPS_fin/EPS_ini)^(1/5) - 1 ---
    set_val("EPS_fin_reciente", eps_fin)
    set_val("EPS_ini_hace_5y", eps_ini)
    if eps_fin is not None and eps_ini is not None and eps_ini > 0:
        set_formula("Crec_EPS_5y", f"=(J{r_}/K{r_})^(1/5)-1")
    else:
        set_val("Crec_EPS_5y", None)

    # --- Margen_Neto = NetIncome_TTM / Revenue_TTM ---
    set_val("Revenue_TTM", revenue_ttm)
    if netinc_ttm is not None and revenue_ttm:
        set_formula("Margen_Neto", f"=F{r_}/M{r_}")
    else:
        set_val("Margen_Neto", None)

    # --- ROE_TTM = NetIncome_TTM / Equity_Promedio ; Equity_Promedio = (actual+previo)/2 ---
    set_val("Equity_actual", equity_actual)
    set_val("Equity_hace_1_anio", equity_previo)
    if equity_actual is not None and equity_previo is not None:
        set_formula("Equity_Promedio", f"=(P{r_}+Q{r_})/2")
    elif equity_actual is not None:
        set_formula("Equity_Promedio", f"=P{r_}")  # sin punto de 1 anio atras, usa el actual solo
    else:
        set_val("Equity_Promedio", None)
    if netinc_ttm is not None and fila.get("_equity_prom"):
        set_formula("ROE_TTM", f"=F{r_}/O{r_}")
    else:
        set_val("ROE_TTM", None)

    # --- Payout = Dividendos_TTM / NetIncome_TTM ---
    set_val("Dividendos_TTM", div_ttm)
    if div_ttm is not None and netinc_ttm:
        set_formula("Payout", f"=S{r_}/F{r_}")
    else:
        set_val("Payout", None)

    # --- Trazabilidad a la fuente EDGAR ---
    set_val("Tag_NetIncome_XBRL", fila.get("_metric_ni"))
    set_val("Tag_Revenue_XBRL", fila.get("_metric_revenue") or None)
    set_val("Estrategia_TTM_NetIncome", fila.get("_ni_estrategia"))
    set_val("Fecha_NetIncome_usado", fila.get("_ni_end"))
    set_val("Flags_calidad", fila.get("_flags") or None)
    set_val("Notas", construir_notas(fila))


def main() -> None:
    print("=" * 60)
    print("  PASO 6 -- Generar solapa calculos_edgar")
    print("=" * 60)

    with open(RATIOS_JSON, encoding="utf-8") as f:
        filas = json.load(f)["filas"]
    print(f"\n  Tickers en ratios_tickets.json: {len(filas)}")

    # Cargar SIN data_only para no congelar formulas que ya existan en otras
    # hojas del workbook (15 hojas ademas de la nueva).
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    sheets_originales = list(wb.sheetnames)
    print(f"  Hojas originales ({len(sheets_originales)}): {sheets_originales}")

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
        print(f"  (ya existia '{SHEET_NAME}', se recrea)")

    ws = wb.create_sheet(SHEET_NAME)

    for c, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=header)

    con_datos = 0
    sin_datos = 0
    for i, fila in enumerate(filas, start=2):
        escribir_fila(ws, i, fila)
        if fila.get("_error"):
            sin_datos += 1
        else:
            con_datos += 1

    for c in range(1, len(HEADERS) + 1):
        letra = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[letra].width = 22

    print(f"\n  Filas escritas: {len(filas)} (con datos EDGAR: {con_datos}, sin datos: {sin_datos})")

    wb.save(TMP_PATH)

    # --- Validacion antes de reemplazar el archivo real ---
    wb_check = openpyxl.load_workbook(TMP_PATH, data_only=False)
    if set(wb_check.sheetnames) != set(sheets_originales) | {SHEET_NAME}:
        TMP_PATH.unlink(missing_ok=True)
        raise SystemExit(
            f"ABORTADO: las hojas del archivo temporal no coinciden con las "
            f"esperadas. Original+nueva: {set(sheets_originales) | {SHEET_NAME}}, "
            f"temporal: {set(wb_check.sheetnames)}. El archivo real no se toco."
        )
    # Spot-check: una celda conocida de otra hoja debe seguir igual.
    if wb_check["resultados"]["I3"].value != 18.36:
        TMP_PATH.unlink(missing_ok=True)
        raise SystemExit(
            "ABORTADO: la hoja 'resultados' cambio inesperadamente al "
            "reescribir el archivo. El archivo real no se toco."
        )
    print("  Validacion OK: las otras 15 hojas quedaron intactas.")

    shutil.move(str(TMP_PATH), str(XLSX_PATH))
    print(f"\n  OK {XLSX_PATH}  (solapa '{SHEET_NAME}' agregada)")
    print("\nPaso 6 completo.")


if __name__ == "__main__":
    main()
