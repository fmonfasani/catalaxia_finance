"""
GRIM (Grimoldi) - PIPELINE COMPLETO v2
Con detección correcta de períodos anuales vs parciales
"""
import csv, re, sqlite3, json, math
from pathlib import Path
from datetime import datetime
import requests, html as ihtml
import yfinance as yf

BASE = Path(r'D:\Software Development\Porfolio\catalaxia-cedears-prod')
H = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120', 'Accept-Language': 'es-AR'}

CODIGOS = {
    "1122500": "Cash", "1121999": "Receivables", "1120100": "Inventory",
    "1139999": "AssetsCurrent", "1110100": "PPE", "1110200": "Intangibles",
    "1119999": "AssetsNonCurrent", "1999999": "Assets",
    "2210999": "Capital", "2211999": "Reservas", "2212999": "ResultadosNoAsignados",
    "2299999": "Equity",
    "2322200": "DebtCurrent", "2321999": "Payables", "2339999": "LiabilitiesCurrent",
    "2312300": "DebtNonCurrent", "2319999": "LiabilitiesNonCurrent", "2399999": "Liabilities",
    "3000100": "Revenue", "3000200": "COGS", "3009999": "GrossProfit",
    "3011600": "DA", "3019999": "OperatingIncome", "3021400": "IngresosFinancieros",
    "3021500": "InterestExpense", "3021800": "RECPAM", "3029999": "PretaxIncome",
    "3031100": "IncomeTax", "3049999": "NetIncome", "3099999": "ResultadoIntegral",
    "3240000": "CashFlowNeto",
    "8000000": "EPS_basico", "8000001": "EPS_diluido", "8000003": "EBIT",
    "8000004": "EBITDA", "8000005": "WorkingCapital",
}

def parse(html_text):
    txt = ihtml.unescape(re.sub("<[^>]+>", " ", html_text))
    txt = re.sub(r"\s+", " ", txt)
    pares = {}
    for m in re.finditer(r"([1-8]\d{6})\s+[A-ZÁÉÍÓÚÑ()/.,\s]+?\s+(-?\d+\.\d{2})", txt):
        pares.setdefault(m.group(1), float(m.group(2)))
    return {CODIGOS[c]: v for c, v in pares.items() if c in CODIGOS}

# ── FASE 1: Discover all NIIF GUIDs for GRIM ──
print("=" * 70)
print("GRIM (Grimoldi SA) - PIPELINE COMPLETO v2")
print("=" * 70)

refined = BASE / 'scripts' / 'tickets' / 'cnv' / 'datos' / 'links_eeff_refined.csv'
grim_niif = []
with open(refined, encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        if row['ticker'].upper() == 'GRIM' and row['formTypeId'] == '147':
            grim_niif.append(row)
grim_niif.sort(key=lambda x: int(float(x['presentationId'])), reverse=True)
print(f"FASE 1: {len(grim_niif)} presentaciones NIIF")

# ── FASE 2: Extract and classify by period type ──
print(f"\nFASE 2: Extrayendo datos...")
all_records = []
for row in grim_niif:
    pid = int(float(row['presentationId']))
    try:
        r = requests.get(row['url'], headers=H, timeout=25, verify=False)
        datos = parse(r.text)
        if 'Assets' in datos and 'Equity' in datos:
            rev = datos.get('Revenue', 0)
            # Classify as annual (Rev > 100B) or partial (Rev < 100B)
            period_type = 'A' if rev > 100_000_000_000 else 'P'
            all_records.append({'pid': pid, 'datos': datos, 'type': period_type, 'rev': rev})
    except:
        pass

# Sort by PID
all_records.sort(key=lambda x: x['pid'])

# Group: separate annual from partial, then take latest of each
annuals = [r for r in all_records if r['type'] == 'A']
partials = [r for r in all_records if r['type'] == 'P']

print(f"  Anuales: {len(annuals)}, Parciales: {len(partials)}")
for r in annuals:
    d = r['datos']
    print(f"    PID={r['pid']}: Rev={d.get('Revenue',0):,.0f} NI={d.get('NetIncome',0):,.0f} Eq={d.get('Equity',0):,.0f}")

# Use the 3 most recent annuals for multi-year comparison
# Assign approximate years (count back from current)
annuals_sorted = sorted(annuals, key=lambda x: x['pid'])
n_annuals = len(annuals_sorted)

# Latest available year from annual data
# We'll assign: latest = current_year-1 (2025), then count backwards
current_year = datetime.now().year  # 2026
latest_annual_yr = current_year - 1  # 2025, since 2026 annual not filed yet

data_by_year = {}
for i, rec in enumerate(annuals_sorted[-5:]):  # Last 5 annuals
    yr = latest_annual_yr - (len(annuals_sorted[-5:]) - 1 - i)
    data_by_year[yr] = rec['datos']

print(f"\n  Años asignados:")
for yr, d in data_by_year.items():
    ident = abs(d.get('Liabilities',0) + d.get('Equity',0) - d.get('Assets',1)) / abs(d.get('Assets',1)) * 100
    print(f"    {yr}: Rev={d.get('Revenue',0):,.0f} NI={d.get('NetIncome',0):,.0f} EBITDA={d.get('EBITDA',0):,.0f} DebtC={d.get('DebtCurrent',0):,.0f} Ident={ident:.2f}%")

# Latest = highest year in dict
latest_yr = max(data_by_year.keys())
latest_d = data_by_year[latest_yr]

# ── FASE 3: Price --  
print(f"\nFASE 3: Precio")
t = yf.Ticker("GRIM.BA")
info = t.info
price_ars = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
year_high = info.get('fiftyTwoWeekHigh')
year_low = info.get('fiftyTwoWeekLow')
shares = info.get('sharesOutstanding')
if not year_high or not year_low:
    hist = t.history(period='1y')
    if len(hist) > 0:
        year_high = year_high or float(hist['High'].max())
        year_low = year_low or float(hist['Low'].min())
try:
    fx_r = requests.get('https://api.bluelytics.com.ar/v2/latest', timeout=10)
    fx = float(fx_r.json().get('blue', {}).get('value_sell', 1200))
except:
    fx = 1200
price_usd = price_ars / fx if price_ars else None
print(f"  Price: ARS${price_ars:,.2f} = USD${price_usd:.2f}")

# ── FASE 4: Ratios ──
print(f"\nFASE 4: 13 ratios (último anual: {latest_yr})")
def sd(a, b): return a / b if (a is not None and b not in (None, 0)) else None

R = {}
R['Precio_u$s'] = price_usd

eps_val = sd(latest_d.get('NetIncome'), shares) if shares else None
R['PER'] = sd(price_ars, eps_val) if eps_val else None
R['EPS'] = eps_val

R['Precio_Max_52s'] = year_high / fx if year_high else None
R['Dif_Max'] = sd(price_ars - year_high, year_high) if year_high else None
R['Precio_Min_52s'] = year_low / fx if year_low else None
R['Dif_Min'] = sd(price_ars - year_low, year_low) if year_low else None

total_debt = (latest_d.get('DebtCurrent', 0) or 0) + (latest_d.get('DebtNonCurrent', 0) or 0)
ebitda_val = latest_d.get('EBITDA', 0)
R['Deuda_EBITDA'] = sd(total_debt, ebitda_val) if ebitda_val else None
R['Margen_Neto'] = sd(latest_d.get('NetIncome'), latest_d.get('Revenue'))
R['ROE'] = sd(latest_d.get('NetIncome'), latest_d.get('Equity'))

# EPS CAGR (up to 5 years)
eps_list = []
for yr in sorted(data_by_year.keys()):
    ni = data_by_year[yr].get('NetIncome')
    if ni is not None and shares:
        eps_list.append((yr, ni / shares))
if len(eps_list) >= 2:
    e0 = eps_list[0][1]
    en = eps_list[-1][1]
    ny = eps_list[-1][0] - eps_list[0][0]
    if e0 > 0 and en > 0 and ny > 0:
        R['Crec_EPS_5y'] = (en / e0) ** (1 / ny) - 1
    elif e0 < 0 and en > 0:
        R['Crec_EPS_5y'] = None
    else:
        R['Crec_EPS_5y'] = sd(en - e0, abs(e0)) if e0 != 0 else None
else:
    R['Crec_EPS_5y'] = None

da = latest_d.get('DA', 0)
fcf = latest_d.get('NetIncome', 0) + da
ce = (latest_d.get('Equity', 0) or 0) + total_debt
R['FCF_CE'] = sd(fcf, ce) if ce else None
R['Payout'] = None

# Print
for label, key, fmt in [
    ('Precio en u$s', 'Precio_u$s', '${:,.2f}'),
    ('PER', 'PER', '{:,.2f}x'),
    ('Máx 52 sem (USD)', 'Precio_Max_52s', '${:,.2f}'),
    ('Dif. Máx', 'Dif_Max', '{:,.2%}'),
    ('Mín 52 sem (USD)', 'Precio_Min_52s', '${:,.2f}'),
    ('Dif. Mín', 'Dif_Min', '{:,.2%}'),
    ('Deuda/EBITDA', 'Deuda_EBITDA', '{:,.2f}x'),
    ('EPS (ARS)', 'EPS', '${:,.2f}'),
    ('Crec. EPS 5y', 'Crec_EPS_5y', '{:,.2%}'),
    ('Margen Neto', 'Margen_Neto', '{:,.2%}'),
    ('ROE', 'ROE', '{:,.2%}'),
    ('FCF/CE', 'FCF_CE', '{:,.2%}'),
    ('Payout', 'Payout', '{}'),
]:
    v = R.get(key)
    if v is not None:
        print(f'  {label:<25} = {fmt.format(v)}')
    else:
        print(f'  {label:<25} = N/A')

# ── FASE 5: Excel Dashboard ──
print(f"\nFASE 5: Dashboard Excel...")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "GRIM Dashboard"

hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
df = Font(name='Calibri', size=11)
tf = Font(name='Calibri', bold=True, size=14)
sf = Font(name='Calibri', bold=True, size=12, color='2F5496')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

ws.merge_cells('A1:O1')
ws['A1'] = 'Dashboard GRIM - Grimoldi SA'
ws['A1'].font = tf
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:O2')
ws['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")} | EEFF anual al {latest_yr} | ARS/USD: {fx}'
ws['A2'].font = Font(name='Calibri', size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

headers = ['Especie','Listada','Precio u$s','PER','Máx 52 sem','Dif. Máx',
           'Mín 52 sem','Dif. Mín','Deuda/EBITDA','EPS (ARS)','Crec. EPS 5y',
           'Margen Neto','ROE','FCF/CE','Payout']

row = 4
for col, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = tb

for i, w in enumerate([8,10,12,8,14,10,14,10,12,12,14,12,10,10,10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def sc(r, c, v, f=None):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v; cell.font = df
        if f: cell.number_format = f
    else:
        cell.value = 'N/A'; cell.font = Font(name='Calibri', size=11, color='999999')
    cell.border = tb; cell.alignment = Alignment(horizontal='center')

r = 5
sc(r, 1, 'GRIM'); sc(r, 2, 'BYMA')
sc(r, 3, R.get('Precio_u$s'), '$#,##0.00')
sc(r, 4, R.get('PER'), '#,##0.00')
sc(r, 5, R.get('Precio_Max_52s'), '$#,##0.00')
sc(r, 6, R.get('Dif_Max'), '0.00%')
sc(r, 7, R.get('Precio_Min_52s'), '$#,##0.00')
sc(r, 8, R.get('Dif_Min'), '0.00%')
sc(r, 9, R.get('Deuda_EBITDA'), '#,##0.00')
sc(r, 10, R.get('EPS'), '#,##0.00')
sc(r, 11, R.get('Crec_EPS_5y'), '0.00%')
sc(r, 12, R.get('Margen_Neto'), '0.00%')
sc(r, 13, R.get('ROE'), '0.00%')
sc(r, 14, R.get('FCF_CE'), '0.00%')
sc(r, 15, R.get('Payout'))

# Financial details
row = 7
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'Detalle Financiero (ARS, último ejercicio)'
ws[f'A{row}'].font = sf
ws[f'A{row}'].fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

r = 8
for col, h in enumerate(['Concepto', 'Valor', 'Concepto', 'Valor'], 1):
    c = ws.cell(row=r, column=col, value=h)
    c.font = hf; c.fill = hfill; c.border = tb

items = [('Revenue', 'Ingresos'), ('NetIncome', 'Resultado Neto'),
         ('EBITDA', 'EBITDA'), ('Equity', 'Patrimonio Neto'),
         ('Assets', 'Activo Total'), ('DebtCurrent', 'Deuda CP'),
         ('DebtNonCurrent', 'Deuda LP'), ('Cash', 'Caja'),
         ('GrossProfit', 'Ganancia Bruta'), ('OperatingIncome', 'Resultado Operativo')]
for i, (k, label) in enumerate(items):
    rr = r + 1 + i
    ws.cell(row=rr, column=1, value=label).font = df; ws.cell(row=rr, column=1).border = tb
    v = latest_d.get(k)
    if v is not None:
        ws.cell(row=rr, column=2, value=v).font = df; ws.cell(row=rr, column=2).number_format = '#,##0'
    else:
        ws.cell(row=rr, column=2, value='N/A').font = Font(color='999999')
    ws.cell(row=rr, column=2).border = tb

# Multi-year history
row = 20
ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Histórico Anual (ARS)'
ws[f'A{row}'].font = sf
ws[f'A{row}'].fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

r = 21
for col, h in enumerate(['Año', 'Revenue', 'NetIncome', 'Equity', 'EBITDA'], 1):
    c = ws.cell(row=r, column=col, value=h)
    c.font = hf; c.fill = hfill; c.border = tb

for i, yr in enumerate(sorted(data_by_year.keys())):
    rr = r + 1 + i
    d = data_by_year[yr]
    ws.cell(row=rr, column=1, value=yr).font = df; ws.cell(row=rr, column=1).border = tb
    for j, k in enumerate(['Revenue', 'NetIncome', 'Equity', 'EBITDA'], 2):
        ws.cell(row=rr, column=j, value=d.get(k, 0)).font = df
        ws.cell(row=rr, column=j).number_format = '#,##0'
        ws.cell(row=rr, column=j).border = tb

# Price section
row = 28
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'Precio y Mercado'
ws[f'A{row}'].font = sf
ws[f'A{row}'].fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

r = 29
for col, h in enumerate(['Concepto', 'Valor', 'Concepto', 'Valor'], 1):
    c = ws.cell(row=r, column=col, value=h)
    c.font = hf; c.fill = hfill; c.border = tb

prices = [('Precio ARS', price_ars), ('Precio USD', price_usd),
          ('52w High ARS', year_high), ('52w Low ARS', year_low),
          ('Shares', shares), ('ARS/USD', fx)]
for i, (label, val) in enumerate(prices):
    rr = r + 1 + i
    ws.cell(row=rr, column=1, value=label).font = df; ws.cell(row=rr, column=1).border = tb
    if val is not None:
        ws.cell(row=rr, column=2, value=val).font = df
        ws.cell(row=rr, column=2).number_format = '#,##0' if abs(val) > 10 else '#,##0.00'
    else:
        ws.cell(row=rr, column=2, value='N/A').font = Font(color='999999')
    ws.cell(row=rr, column=2).border = tb

output = BASE / 'Dashboard_GRIM_completo.xlsx'
wb.save(output)
print(f"\n{'='*70}")
print(f"DASHBOARD: {output}")
print(f"{'='*70}")
print(f"\nRESUMEN FINAL GRIM (Grimoldi SA):")
print(f"  Precio: USD${R.get('Precio_u$s',0):.2f} (ARS${price_ars:,.0f})")
print(f"  PER: {R.get('PER', 'N/A')}")
print(f"  Margen Neto: {R.get('Margen_Neto', 'N/A')}")
print(f"  ROE: {R.get('ROE', 'N/A')}")
print(f"  Deuda/EBITDA: {R.get('Deuda_EBITDA', 'N/A')}")
print(f"  EPS: {R.get('EPS', 'N/A')}")
print(f"  Crec. EPS: {R.get('Crec_EPS_5y', 'N/A')}")
print(f"  FCF/CE: {R.get('FCF_CE', 'N/A')}")
print(f"  Años usados: {sorted(data_by_year.keys())}")
